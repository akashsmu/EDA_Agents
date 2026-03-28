import pytest

def test_pandas_functionality():
    try:
        import pandas as pd
        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        assert not df.empty
        assert list(df.columns) == ["A", "B"]
        assert df["A"].sum() == 3
    except ImportError:
        pytest.fail("pandas library is missing")
    except Exception as e:
        pytest.fail(f"pandas basic functionality failed: {e}")

def test_langchain_functionality():
    try:
        from langchain_core.messages import HumanMessage, AIMessage
        msg = HumanMessage(content="Hello")
        assert msg.content == "Hello"
    except ImportError:
        pytest.fail("langchain library or core components missing")
    except Exception as e:
        pytest.fail(f"langchain basic functionality failed: {e}")

def test_langgraph_functionality():
    try:
        from langgraph.graph import StateGraph, END
        from typing import TypedDict

        class State(TypedDict):
            test_val: str
        
        graph = StateGraph(State)
        graph.add_node("start", lambda state: {"test_val": "done"})
        graph.set_entry_point("start")
        graph.add_edge("start", END)
        app = graph.compile()
        assert app is not None
    except ImportError:
        pytest.fail("langgraph library is missing")
    except Exception as e:
        pytest.fail(f"langgraph basic functionality failed: {e}")

def test_openpyxl_functionality():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        assert ws["A1"].value == "Test"
    except ImportError:
        pytest.fail("openpyxl library is missing")
    except Exception as e:
        pytest.fail(f"openpyxl basic functionality failed: {e}")

def test_plotly_functionality():
    try:
        import plotly.graph_objects as go
        fig = go.Figure(data=go.Bar(y=[2, 3, 1]))
        assert fig is not None
        assert hasattr(fig, "show")
    except ImportError:
        pytest.fail("plotly library is missing")
    except Exception as e:
        pytest.fail(f"plotly basic functionality failed: {e}")

def test_streamlit_functionality():
    try:
        import streamlit as st
        # Just testing that the streamlit object has the expected core properties
        assert hasattr(st, "write")
        assert hasattr(st, "dataframe")
    except ImportError:
        pytest.fail("streamlit library is missing")
    except Exception as e:
        pytest.fail(f"streamlit basic functionality failed: {e}")
